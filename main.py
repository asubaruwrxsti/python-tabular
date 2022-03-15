import http.server  # Our http server handler for http requests
import socketserver  # Establish the TCP Socket connections

PORT = 8000
out = open("index.html", "w")  # file open

from openpyxl import *

wb = load_workbook('sampledata.xlsx', data_only=True)
ws = wb.active

out.write("<!DOCTYPE html> <head> <style> \n")  # html head & style opener
out.write("table {font-family: arial, sans-serif; border-collapse: collapse; width: 70%;} \n")  # table css
out.write("td, th { border: 1px solid #dddddd; text-align: center; padding: 3px;} \n")  # table data / header css
out.write("tr:nth-child(even) {background-color: #dddddd;}\n")  # table row css
out.write("div{padding-top: 50px;}")
out.write("</style>\n")  # head & style closer

out.write('<script> \n function button_switch(i){ \n switch(i){ \n')

# loop to add data to table
j = 0
for name in wb.sheetnames:
    ws = wb[name]
    out.write('  case {:}: \n   var element = document.getElementById("table_data"); \n   element.innerHTML= "'.format(j))
    for value in ws.values:
        out.write("<tr>")
        for i in range(len(value)):
            out.write("<td> {:} </td>".format(value[i]))
        out.write("</tr>")
    out.write('";\n   break; \n')
    j += 1

# } closer, close case of loop
out.write(' }\n}\n')
out.write('function dropdown_menu() { \n var list = document.getElementById("list"); \n button_switch(list.selectedIndex);}\n')
out.write('</script></head><body onLoad="javascript:button_switch(0)">\n')
out.write('<select id="list">')

# loop to add all buttons
i = 0
for name in wb.sheetnames:
    out.write('<option>{:}</option>\n'.format(name))
    i += 1
out.write('</select>')
out.write('<button onclick="dropdown_menu()">OK</button>')
out.write('\n<div><table id="table_data"></table></div>')

out.write("</body></html>")  # html closer
out.close()  # file close


# server setup
class HttpHandler(http.server.SimpleHTTPRequestHandler):
    def do_GET(self):
        self.path = 'index.html'
        return http.server.SimpleHTTPRequestHandler.do_GET(self)


Handler = HttpHandler

with socketserver.TCPServer(("", PORT), Handler) as httpd:
    print("Http Server Serving at port", PORT)
    httpd.serve_forever()
